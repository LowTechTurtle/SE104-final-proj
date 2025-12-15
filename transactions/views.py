# Standard library imports
import json
import logging

# Django core imports
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.shortcuts import render
from django.db import transaction

# Class-based views
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView

# Authentication and permissions
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

# Third-party packages
from openpyxl import Workbook

# Local app imports
from store.models import Item
from accounts.models import Customer
from .models import Sale, Purchase, SaleDetail
from .forms import PurchaseForm
import openpyxl

from django.contrib.auth.decorators import login_required

logger = logging.getLogger(__name__)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


def export_sales_to_excel(request):
    # Create a workbook and select the active worksheet.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sales'

    # Define the column headers
    columns = [
        'ID', 'Date', 'Customer', 'Sub Total',
        'Grand Total', 'Tax Amount', 'Tax Percentage',
        'Amount Paid', 'Amount Change'
    ]
    worksheet.append(columns)

    # Fetch sales data
    sales = Sale.objects.all()

    for sale in sales:
        # Convert timezone-aware datetime to naive datetime
        if sale.date_added.tzinfo is not None:
            date_added = sale.date_added.replace(tzinfo=None)
        else:
            date_added = sale.date_added

        worksheet.append([
            sale.id,
            date_added,
            sale.customer.phone,
            sale.sub_total,
            sale.grand_total,
            sale.tax_amount,
            sale.tax_percentage,
            sale.amount_paid,
            sale.amount_change
        ])

    # Set up the response to send the file
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = 'attachment; filename=sales.xlsx'
    workbook.save(response)

    return response


@login_required
def export_purchases(request):
    """
    Export danh sách Purchase Order an toàn (Fix lỗi NoneType)
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Purchase_Orders.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"

    # 1. Header
    columns = [
        'ID', 
        'Date', 
        'Vendor',          # Cột gây lỗi
        'Item Name', 
        'Quantity', 
        'Price (Ksh)', 
        'Total Value', 
        'Status',
        'Delivery Date'
    ]
    ws.append(columns)

    # Format Header
    header_font = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # 2. Query dữ liệu
    rows = Purchase.objects.all().select_related('vendor', 'item').order_by('-order_date')

    for p in rows:
        # -- Xử lý Ngày tháng --
        o_date = p.order_date.strftime('%Y-%m-%d %H:%M') if p.order_date else "-"
        d_date = p.delivery_date.strftime('%Y-%m-%d') if p.delivery_date else "-"

        # -- Xử lý Vendor (FIX LỖI NONE TYPE TẠI ĐÂY) --
        # Kiểm tra: Nếu có vendor thì lấy tên, nếu không thì ghi 'N/A'
        if p.vendor:
            vendor_name = p.vendor.name
        else:
            vendor_name = "N/A" 

        # -- Xử lý Item --
        item_name = p.item.name if p.item else "Unknown"

        # -- Xử lý Status (Hiển thị chữ thay vì ký tự P/S) --
        status_display = "Pending"
        if p.delivery_status == 'S':
            status_display = "Successful"

        ws.append([
            p.id,
            o_date,
            vendor_name,    # Biến đã được xử lý an toàn
            item_name,
            p.quantity,
            p.price,
            p.total_value,
            status_display,
            d_date
        ])

    # 3. Auto-fit cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


class SaleListView(LoginRequiredMixin, ListView):
    """
    View to list all sales with pagination.
    """

    model = Sale
    template_name = "transactions/sales_list.html"
    context_object_name = "sales"
    paginate_by = 10

    ordering = ['-date_added']


class SaleDetailView(LoginRequiredMixin, DetailView):
    """
    View to display details of a specific sale.
    """

    model = Sale
    template_name = "transactions/saledetail.html"

def SaleCreateView(request):
    context = {
        "active_icon": "sales",
        "customers": [c.to_select2() for c in Customer.objects.all()]
    }

    if request.method == 'POST':
        if is_ajax(request=request):
            try:
                # Load the JSON data from the request body
                data = json.loads(request.body)
                logger.info(f"Received data: {data}")

                # Validate required fields
                required_fields = [
                    'customer', 'sub_total', 'grand_total',
                    'amount_paid', 'amount_change', 'items'
                ]
                for field in required_fields:
                    if field not in data:
                        raise ValueError(f"Missing required field: {field}")

                # --- ĐOẠN ĐÃ SỬA (FIX BUG SỐ 1) ---
                # Lấy ID khách hàng từ dữ liệu
                customer_id = data.get('customer')
                customer_obj = None  # Mặc định là None (Khách ẩn danh)

                # Kiểm tra: nếu customer_id không rỗng và không phải None thì mới đi tìm trong DB
                if customer_id and str(customer_id).strip():
                    customer_obj = Customer.objects.get(id=int(customer_id))
                # ----------------------------------

                # Create sale attributes
                sale_attributes = {
                    "customer": customer_obj, # Truyền object khách hàng (hoặc None) vào đây
                    "sub_total": float(data["sub_total"]),
                    "grand_total": float(data["grand_total"]),
                    "tax_amount": float(data.get("tax_amount", 0.0)),
                    "tax_percentage": float(data.get("tax_percentage", 0.0)),
                    "amount_paid": float(data["amount_paid"]),
                    "amount_change": float(data["amount_change"]),
                }

                # Use a transaction to ensure atomicity
                with transaction.atomic():
                    # Create the sale
                    new_sale = Sale.objects.create(**sale_attributes)
                    logger.info(f"Sale created: {new_sale}")

                    # Create sale details and update item quantities
                    items = data["items"]
                    if not isinstance(items, list):
                        raise ValueError("Items should be a list")

                    for item in items:
                        if not all(
                            k in item for k in [
                                "id", "price", "quantity", "total_item"
                            ]
                        ):
                            raise ValueError("Item is missing required fields")

                        item_instance = Item.objects.get(id=int(item["id"]))
                        if item_instance.quantity < int(item["quantity"]):
                            raise ValueError(f"Not enough stock for item: {item_instance.name}")

                        detail_attributes = {
                            "sale": new_sale,
                            "item": item_instance,
                            "price": float(item["price"]),
                            "quantity": int(item["quantity"]),
                            "total_detail": float(item["total_item"])
                        }
                        SaleDetail.objects.create(**detail_attributes)
                        logger.info(f"Sale detail created: {detail_attributes}")

                        # Reduce item quantity
                        item_instance.quantity -= int(item["quantity"])
                        item_instance.save()

                return JsonResponse(
                    {
                        'status': 'success',
                        'message': 'Sale created successfully!',
                        'redirect': '/transactions/sales/'
                    }
                )

            except json.JSONDecodeError:
                return JsonResponse(
                    {
                        'status': 'error',
                        'message': 'Invalid JSON format in request body!'
                    }, status=400)
            except Customer.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Customer does not exist!'
                    }, status=400)
            except Item.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Item does not exist!'
                    }, status=400)
            except ValueError as ve:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Value error: {str(ve)}'
                    }, status=400)
            except TypeError as te:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Type error: {str(te)}'
                    }, status=400)
            except Exception as e:
                logger.error(f"Exception during sale creation: {e}")
                return JsonResponse(
                    {
                        'status': 'error',
                        'message': (
                            f'There was an error during the creation: {str(e)}'
                        )
                    }, status=500)

    return render(request, "transactions/sale_create.html", context=context)

class SaleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    View to delete a sale.
    """

    model = Sale
    template_name = "transactions/saledelete.html"

    def get_success_url(self):
        """
        Redirect to the sales list after successful deletion.
        """
        return reverse("saleslist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser


class PurchaseListView(LoginRequiredMixin, ListView):
    """
    View to list all purchases with pagination.
    """

    model = Purchase
    template_name = "transactions/purchases_list.html"
    context_object_name = "purchases"
    paginate_by = 10


class PurchaseDetailView(LoginRequiredMixin, DetailView):
    """
    View to display details of a specific purchase.
    """

    model = Purchase
    template_name = "transactions/purchasedetail.html"


class PurchaseCreateView(LoginRequiredMixin, CreateView):
    """
    View to create a new purchase.
    """

    model = Purchase
    form_class = PurchaseForm
    template_name = "transactions/purchases_form.html"

    def get_success_url(self):
        """
        Redirect to the purchases list after successful form submission.
        """
        return reverse("purchaseslist")


class PurchaseUpdateView(LoginRequiredMixin, UpdateView):
    """
    View to update an existing purchase.
    """

    model = Purchase
    form_class = PurchaseForm
    template_name = "transactions/purchases_form.html"

    def get_success_url(self):
        """
        Redirect to the purchases list after successful form submission.
        """
        return reverse("purchaseslist")


class PurchaseDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    View to delete a purchase.
    """

    model = Purchase
    template_name = "transactions/purchasedelete.html"

    def get_success_url(self):
        """
        Redirect to the purchases list after successful deletion.
        """
        return reverse("purchaseslist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser
